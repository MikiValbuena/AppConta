"""
Carga de datos desde Conta.xlsx para el frontend.
Lee la hoja del ano 2025 y estructura los datos.
"""

import openpyxl
from datetime import datetime
from collections import defaultdict


class DataLoader:
    """Carga y estructura los datos desde el Excel."""

    def __init__(self, ruta_excel="Conta.xlsx", hoja="2025"):
        self.ruta_excel = ruta_excel
        self.hoja = hoja
        self.movimientos = []
        self.categorias = set()
        self.subcategorias = set()
        self.balance_mensual = {}
        self.saldo_inicial = {}
        self.cuentas = {}
        self._cargar()

    def _cargar(self):
        """Carga los datos de la hoja 2025."""
        wb = openpyxl.load_workbook(self.ruta_excel, data_only=True)
        ws = wb[self.hoja]

        # Leer estructura de cuentas desde las filas de encabezado
        # Fila 2-3: Informacion de cuentas
        # Filas 4-7: Saldos de arrastre
        for row in ws.iter_rows(min_row=2, max_row=3, values_only=False):
            for c in row:
                if c.value and isinstance(c.value, str):
                    pass  # Info de cuentas

        # Las cuentas se identifican por las columnas:
        # H-I: OpenBank Principal, J-K: OpenBank Socu, L-M: Efectivo
        # N-O: Interactive Broker, P-Q: Degiro
        self.cuentas = {
            "OpenBank Principal": {"col_ingreso": 7, "col_gasto": 8},   # H, I (0-indexed)
            "OpenBank Socu":      {"col_ingreso": 9, "col_gasto": 10},   # J, K
            "Efectivo":           {"col_ingreso": 11, "col_gasto": 12},  # L, M
            "IBKR":               {"col_ingreso": 13, "col_gasto": 14},  # N, O
            "Degiro":             {"col_ingreso": 15, "col_gasto": 16},  # P, Q
        }

        # Leer datos desde fila 9 (despues de encabezados)
        # Estructura: B=Año, C=Mes, D=Fecha, E=Cat1, F=Cat2, G=Cat3,
        #             H=Ingreso, I=Gasto (OpenBank Principal)
        #             J=Ingreso, K=Gasto (OpenBank Socu)
        #             L=Ingreso, M=Gasto (Efectivo)
        #             N=Invertido, O=Retorno (IBKR)
        #             P=Invertido, Q=Retorno (Degiro)
        for row in ws.iter_rows(min_row=9, max_row=ws.max_row, values_only=True):
            fecha = row[3]  # Col D (0-indexed: 3)
            cat1 = row[4]   # Col E
            cat2 = row[5]   # Col F
            cat3 = row[6]   # Col G

            # Detectar filas de totales (Saldo, Total Mes, Balance Mes, etc.)
            if not fecha or not isinstance(fecha, (datetime, str)):
                continue

            if isinstance(cat1, str) and cat1 in (
                "Total Ingreso Mes", "Total Gasto Mes", "Balance Mes",
                "Saldo Mes", "Saldo Final Mes", "Saldo Total Anual",
                "Resultado Anual", "Saldo Anual", "Acciones", "Cuentas",
                "Saldo Año Ant."
            ):
                # Es una fila de totales, no un movimiento
                if cat1 == "Total Ingreso Mes":
                    self._procesar_fila_total(row, "ingreso")
                elif cat1 == "Total Gasto Mes":
                    self._procesar_fila_total(row, "gasto")
                elif cat1 == "Balance Mes":
                    self._procesar_fila_balance(row)
                continue

            if not cat1 or not isinstance(cat1, str):
                continue

            if cat1.strip() == "":
                continue

            self.categorias.add(cat1)
            if cat2 and isinstance(cat2, str):
                self.subcategorias.add(cat2)

            # Procesar cada cuenta
            self._procesar_movimiento(fecha, cat1, cat2, cat3, row)

        wb.close()
        self._calcular_balance_mensual()

    def _procesar_movimiento(self, fecha, cat1, cat2, cat3, row):
        """Procesa una fila de movimiento."""
        for nombre_cuenta, cols in self.cuentas.items():
            ingreso = row[cols["col_ingreso"]]
            gasto = row[cols["col_gasto"]]

            if ingreso is not None and isinstance(ingreso, (int, float)) and ingreso != 0:
                self.movimientos.append({
                    "fecha": fecha,
                    "categoria": cat1,
                    "subcategoria": cat2 or "",
                    "detalle": cat3 or "",
                    "cuenta": nombre_cuenta,
                    "tipo": "ingreso",
                    "importe": float(ingreso),
                })

            if gasto is not None and isinstance(gasto, (int, float)) and gasto != 0:
                self.movimientos.append({
                    "fecha": fecha,
                    "categoria": cat1,
                    "subcategoria": cat2 or "",
                    "detalle": cat3 or "",
                    "cuenta": nombre_cuenta,
                    "tipo": "gasto",
                    "importe": float(gasto),
                })

    def _procesar_fila_total(self, row, tipo):
        """Procesa una fila de totales mensuales."""
        mes = row[2] if row[2] else "Total"
        if isinstance(mes, datetime):
            mes = mes.strftime("%B")
        # Las columnas de total son las mismas
        for nombre_cuenta, cols in self.cuentas.items():
            if tipo == "ingreso":
                val = row[cols["col_ingreso"]]
            else:
                val = row[cols["col_gasto"]]
            if val is not None and isinstance(val, (int, float)):
                pass  # Podriamos almacenar totales si hiciera falta

    def _procesar_fila_balance(self, row):
        """Procesa una fila de balance mensual."""
        pass

    def _calcular_balance_mensual(self):
        """Calcula balance mensual por cuenta."""
        for movimiento in self.movimientos:
            fecha = movimiento["fecha"]
            if isinstance(fecha, datetime):
                mes = fecha.month
                ano = fecha.year
            else:
                continue

            clave = f"{ano}-{mes:02d}"
            if clave not in self.balance_mensual:
                self.balance_mensual[clave] = {
                    "ingresos": defaultdict(float),
                    "gastos": defaultdict(float),
                    "balance": defaultdict(float),
                }

            cuenta = movimiento["cuenta"]
            if movimiento["tipo"] == "ingreso":
                self.balance_mensual[clave]["ingresos"][cuenta] += movimiento["importe"]
            else:
                self.balance_mensual[clave]["gastos"][cuenta] += movimiento["importe"]

            total_ing = sum(self.balance_mensual[clave]["ingresos"].values())
            total_gas = sum(self.balance_mensual[clave]["gastos"].values())
            self.balance_mensual[clave]["balance"]["Total"] = total_ing - total_gas

    def get_movimientos(self, mes=None, categoria=None):
        """Filtra movimientos por mes y/o categoria."""
        resultado = self.movimientos
        if mes:
            resultado = [m for m in resultado if self._fecha_a_mes(m["fecha"]) == mes]
        if categoria:
            resultado = [m for m in resultado if m["categoria"] == categoria]
        return sorted(resultado, key=lambda x: x["fecha"] if isinstance(x["fecha"], datetime) else datetime.min, reverse=True)

    def get_gasto_por_categoria(self):
        """Agrupa gastos por categoria."""
        gastos = defaultdict(float)
        for m in self.movimientos:
            if m["tipo"] == "gasto":
                gastos[m["categoria"]] += m["importe"]
        return dict(sorted(gastos.items(), key=lambda x: x[1], reverse=True))

    def get_ingreso_por_categoria(self):
        """Agrupa ingresos por categoria."""
        ingresos = defaultdict(float)
        for m in self.movimientos:
            if m["tipo"] == "ingreso":
                ingresos[m["categoria"]] += m["importe"]
        return dict(sorted(ingresos.items(), key=lambda x: x[1], reverse=True))

    def get_resumen_anual(self):
        """Resumen anual: total ingresos, gastos, balance."""
        total_ingresos = sum(m["importe"] for m in self.movimientos if m["tipo"] == "ingreso")
        total_gastos = sum(m["importe"] for m in self.movimientos if m["tipo"] == "gasto")
        return {
            "ingresos": total_ingresos,
            "gastos": total_gastos,
            "balance": total_ingresos - total_gastos,
            "total_movimientos": len(self.movimientos),
        }

    def get_meses_disponibles(self):
        """Lista de meses con datos."""
        meses = set()
        for m in self.movimientos:
            if isinstance(m["fecha"], datetime):
                meses.add(m["fecha"].month)
        return sorted(meses)

    @staticmethod
    def _fecha_a_mes(fecha):
        if isinstance(fecha, datetime):
            return fecha.month
        return None

    @staticmethod
    def nombre_mes(numero):
        """Convierte numero de mes a nombre."""
        nombres = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
        }
        return nombres.get(numero, f"Mes {numero}")
