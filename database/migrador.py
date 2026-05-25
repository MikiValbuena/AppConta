"""
Migrador de datos desde Conta.xlsx a SQLite (optimizado).
Importa todos los anos en batch con una sola transaccion.
"""

import openpyxl
from datetime import datetime, date
from database.conexion import DatabaseManager


class Migrador:
    """Importa datos del Excel a SQLite."""

    HOJAS = ["2026", "2025", "2024", "2023", "2022", "2021",
             "2020", "2019", "2018", "2017", "2016"]

    CUENTAS_MAP = {
        "OpenBank Principal": {"col_ingreso": 7, "col_gasto": 8},
        "OpenBank Socu":      {"col_ingreso": 9, "col_gasto": 10},
        "Efectivo":           {"col_ingreso": 11, "col_gasto": 12},
        "IBKR":               {"col_ingreso": 13, "col_gasto": 14},
        "Degiro":             {"col_ingreso": 15, "col_gasto": 16},
    }

    EXCLUIR_CATEGORIAS = {
        "Total Ingreso Mes", "Total Gasto Mes", "Balance Mes",
        "Saldo Mes", "Saldo Final Mes", "Saldo Total Anual",
        "Resultado Anual", "Saldo Anual", "Acciones", "Cuentas",
        "Saldo A\u00f1o Ant.", "Balance A\u00f1o", "Saldo Total Anual Ctas.",
        "Total Op.", "Promedio Op.->",
    }

    def __init__(self, ruta_excel="Conta.xlsx"):
        self.ruta_excel = ruta_excel
        self.db = DatabaseManager()
        self.total_importados = 0
        self.batch = []

    def ejecutar(self, progreso_callback=None):
        """Ejecuta la migracion completa en batch."""
        self.total_importados = 0
        self.batch = []
        wb = openpyxl.load_workbook(self.ruta_excel, data_only=True)

        for hoja in self.HOJAS:
            if hoja not in wb.sheetnames:
                continue
            ws = wb[hoja]
            self._procesar_hoja(ws)

            if progreso_callback:
                progreso_callback(hoja, len(self.batch))

        wb.close()

        # Insertar todo en una sola transaccion
        if self.batch:
            self.db.conexion.execute("BEGIN TRANSACTION")
            self.db.conexion.executemany(
                """INSERT INTO movimientos
                   (fecha, categoria1, categoria2, categoria3, cuenta, tipo, importe)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                self.batch
            )
            self.db.conexion.commit()
            self.total_importados = len(self.batch)

        return self.total_importados

    def _procesar_hoja(self, ws):
        """Procesa una hoja de ano y acumula en batch."""
        for row in ws.iter_rows(min_row=9, max_row=ws.max_row, values_only=True):
            if len(row) < 5:
                continue

            fecha = row[3]
            cat1 = row[4]

            if not fecha or not isinstance(fecha, (datetime, str)):
                continue

            # Normalizar fecha
            if isinstance(fecha, datetime):
                fecha = fecha.date().isoformat()

            if isinstance(cat1, str) and cat1 in self.EXCLUIR_CATEGORIAS:
                continue
            if not cat1 or not isinstance(cat1, str) or cat1.strip() == "":
                continue

            cat2 = row[5] if len(row) > 5 and row[5] else ""
            cat3 = row[6] if len(row) > 6 and row[6] else ""

            if isinstance(cat2, str) and cat2 in self.EXCLUIR_CATEGORIAS:
                cat2 = ""
            if isinstance(cat3, str) and cat3 in self.EXCLUIR_CATEGORIAS:
                cat3 = ""

            self._procesar_fila(fecha, cat1, str(cat2) if cat2 else "",
                                str(cat3) if cat3 else "", row)

    def _procesar_fila(self, fecha, cat1, cat2, cat3, row):
        """Acumula movimientos de una fila en el batch."""
        max_col = len(row)
        for nombre_cuenta, cols in self.CUENTAS_MAP.items():
            if cols["col_ingreso"] >= max_col or cols["col_gasto"] >= max_col:
                continue

            ingreso = row[cols["col_ingreso"]]
            gasto = row[cols["col_gasto"]]

            if ingreso is not None and isinstance(ingreso, (int, float)) and ingreso != 0:
                self.batch.append((fecha, cat1, cat2, cat3, nombre_cuenta, "ingreso", float(ingreso)))

            if gasto is not None and isinstance(gasto, (int, float)) and gasto != 0:
                tipo = "invertido" if nombre_cuenta in ("IBKR", "Degiro") else "gasto"
                self.batch.append((fecha, cat1, cat2, cat3, nombre_cuenta, tipo, float(gasto)))

    def necesita_migracion(self):
        """Verifica si hay datos en la BD."""
        count = self.db.conexion.execute(
            "SELECT COUNT(*) FROM movimientos"
        ).fetchone()[0]
        return count == 0
